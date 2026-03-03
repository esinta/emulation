#!/usr/bin/env python3
"""
============================================================================
ESINTA EMULATION — AUTHORIZED SECURITY TESTING ONLY

This code replicates documented malware behavior for defensive security
testing and endpoint telemetry validation. It does NOT contain real malware
payloads, exploits, or evasion techniques.

LEGAL: Only run this in environments you own or have explicit written
authorization to test. Unauthorized use may violate computer fraud laws.

Final payload: calc.exe (safe, benign)
C2: Local network only (hardcoded private IP: 192.168.0.148)
============================================================================

MuddyCalc Stage 1 — Expense Report Spreadsheet Builder

Generates a realistic-looking expense report spreadsheet that serves as the
initial access document. In a real MuddyWater campaign, the victim receives
a macro-enabled spreadsheet via spearphishing email.

Usage:
    pip install openpyxl
    python3 build_spreadsheet.py

Output:
    Q4_2025_ExpenseReport_FINAL_FINAL_v3_DO_NOT_DELETE.xlsx

After generating the .xlsx, you must manually:
1. Open it in LibreOffice Calc
2. Add the macro from macro.vba (Tools → Macros → paste code)
3. Save As → .xlsm format

MITRE ATT&CK:
    T1566.001 — Phishing: Spearphishing Attachment
    T1204.002 — User Execution: Malicious File
"""

import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.")
    print("Install it with: pip install openpyxl")
    sys.exit(1)


def load_expense_data():
    """Load expense data from JSON file."""
    json_path = os.path.join(os.path.dirname(__file__), "expense_data.json")
    with open(json_path, "r") as f:
        return json.load(f)


def build_spreadsheet():
    """Generate the expense report spreadsheet."""
    data = load_expense_data()
    wb = Workbook()
    ws = wb.active
    ws.title = "Q4 2025 Expenses"

    # --- Styles ---
    header_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="4472C4")
    col_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    even_row_fill = PatternFill(
        start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
    )
    total_font = Font(name="Calibri", size=11, bold=True)
    total_fill = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )
    currency_format = '"$"#,##0.00'
    date_format = "MM/DD/YYYY"
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Company header ---
    ws.merge_cells("A1:E1")
    ws["A1"] = data["company"]
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"{data['report_title']} — {data['submitter']}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # Blank row
    start_row = 4

    # --- Column headers ---
    headers = ["Date", "Category", "Description", "Amount", "Approved By"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # --- Expense data rows ---
    for row_idx, expense in enumerate(data["expenses"]):
        row = start_row + 1 + row_idx
        is_even = row_idx % 2 == 0

        # Date
        date_obj = datetime.strptime(expense["date"], "%Y-%m-%d")
        cell = ws.cell(row=row, column=1, value=date_obj)
        cell.number_format = date_format
        cell.border = thin_border
        if is_even:
            cell.fill = even_row_fill

        # Category
        cell = ws.cell(row=row, column=2, value=expense["category"])
        cell.border = thin_border
        if is_even:
            cell.fill = even_row_fill

        # Description
        cell = ws.cell(row=row, column=3, value=expense["description"])
        cell.border = thin_border
        if is_even:
            cell.fill = even_row_fill

        # Amount
        cell = ws.cell(row=row, column=4, value=expense["amount"])
        cell.number_format = currency_format
        cell.border = thin_border
        if is_even:
            cell.fill = even_row_fill

        # Approved By
        cell = ws.cell(row=row, column=5, value=expense["approved_by"])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        if is_even:
            cell.fill = even_row_fill

    # --- Totals row ---
    total_row = start_row + 1 + len(data["expenses"])
    num_expenses = len(data["expenses"])
    first_amount_row = start_row + 1
    last_amount_row = start_row + num_expenses

    ws.cell(row=total_row, column=1, value="").border = thin_border
    ws.cell(row=total_row, column=2, value="").border = thin_border

    total_label = ws.cell(row=total_row, column=3, value="TOTAL")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.border = thin_border
    total_label.alignment = Alignment(horizontal="right")

    total_amount = ws.cell(
        row=total_row,
        column=4,
        value=f"=SUM(D{first_amount_row}:D{last_amount_row})",
    )
    total_amount.font = total_font
    total_amount.fill = total_fill
    total_amount.number_format = currency_format
    total_amount.border = thin_border

    ws.cell(row=total_row, column=5, value="").fill = total_fill
    ws.cell(row=total_row, column=5).border = thin_border

    # Also fill columns A and B of total row
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=2).fill = total_fill

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # --- Save ---
    output_filename = (
        "Q4_2025_ExpenseReport_FINAL_FINAL_v3_DO_NOT_DELETE.xlsx"
    )
    output_path = os.path.join(os.path.dirname(__file__), output_filename)
    wb.save(output_path)

    print(f"Spreadsheet generated: {output_filename}")
    print()
    print("Next steps:")
    print("  1. Run: python3 encode_payload.py")
    print("     (generates the base64-encoded PowerShell command)")
    print()
    print(f"  2. Open {output_filename} in LibreOffice Calc")
    print("  3. Tools → Macros → Organize Dialogs → BASIC")
    print("     Paste the macro code from the encode_payload.py output")
    print("  4. File → Save As → Microsoft Excel 2007-365 (.xlsm)")
    print(
        "     Name: Q4_2025_ExpenseReport_FINAL_FINAL_v3_DO_NOT_DELETE.xlsm"
    )
    print("  5. Copy .xlsm to muddycalc/ root directory for C2 serving")

    return output_path


if __name__ == "__main__":
    build_spreadsheet()
